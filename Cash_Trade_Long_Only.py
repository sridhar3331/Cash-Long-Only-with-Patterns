import os
working_dir = os.chdir(r"D:\ALGO TRADING\LIVE\ZERODHA CASH")
import warnings
# Suppress future warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import pytz
UTC = pytz.timezone('Asia/Kolkata')
import time

##### Import From Common Functions ######
from Common_Functions import data_list,is_required_time,getting_ohlc,instrument_lookup
#########################################

##### Importing From User Input Library ######
from User_Input import Total_Cash_per_position,tele_msg,Long_max_open_trade,EXIT_TIME,Tickers,RISING_WEDGE_A_LONG_Tickers,RISING_WEDGE_C_LONG_Tickers,RISING_WEDGE_3P_LONG_Tickers
##############################################

##### Importing Streaming LTP DATA #############
from Streaming_Ltp import spot_prices
##############################################


import numpy as np
import pandas as pd
from scipy.signal import argrelextrema
from openpyxl import load_workbook
import datetime as dt
import math
from pyotp import TOTP
from kiteconnect import KiteConnect
import sridharzerodhacred
import urllib
from selenium import webdriver


RISING_WEDGE_A_LONG_targetx = 0.53

RISING_WEDGE_C_LONG_targetx = 0.53

RISING_WEDGE_3P_LONG_targetx = 0.33



long_argrel_window = 7

RISING_WEDGE_A_LONG_result_dict = {}

RISING_WEDGE_C_LONG_result_dict = {}

RISING_WEDGE_3P_LONG_result_dict = {}


# kite connect login

# Declaring Login Credentials in a variable
api_key = sridharzerodhacred.apikey
key_secret = sridharzerodhacred.api_secret
userID = sridharzerodhacred.userid
pwd = sridharzerodhacred.password
totp_key = sridharzerodhacred.totp



# Using Selenium Web driver to get login url
browser = webdriver.Chrome()
browser.get("https://kite.zerodha.com/connect/login?v=3&api_key="+urllib.parse.quote_plus(api_key))
browser.implicitly_wait(5)


# Detecting User Name and Password field in a browser
username = browser.find_element("xpath", '/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[1]/input')
password = browser.find_element("xpath", '/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[2]/input') 

# Sending user name and password to the field in the browser
username.send_keys(userID)
password.send_keys(pwd)

#Checkbox finding
browser.find_element("xpath", '/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[3]/div/label/span[1]').click()

# Click Login Button finding
browser.find_element("xpath", '/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/form/div[4]/button').click()
time.sleep(2)


#Provide OTP programatically
pin = browser.find_element("xpath", '/html/body/div[1]/div/div[2]/div[1]/div[2]/div/div[2]/form/div[1]/input')
totp = TOTP(totp_key)
token = totp.now()
pin.send_keys(token)

#Click button and Get Token
#browser.find_element("xpath", '/html/body/div[1]/div/div[2]/div[1]/div[2]/div/div[2]/form/div[2]/button').click()
time.sleep(1)


temp_token=browser.current_url.split('request_token=')[1][:32]
# Save in Database or text File
print('temp_token', temp_token)

#Generate Session

kite = KiteConnect(api_key=api_key)

data = kite.generate_session(temp_token, api_secret=key_secret)
kite.set_access_token(data["access_token"])
browser.close()

# Getting Instrument for Zerodha
nse_futdump = kite.instruments("NFO")
nse_fut = pd.DataFrame(nse_futdump)
nse_fut = nse_fut[nse_fut.instrument_type=='FUT']



# 30 min close
def min_30_close(symbol):
    
    closevalue = data_list[symbol]['Close'].iloc[-2]
    
    
    return closevalue


def get_min_max(data, argrel_window):

    # Use the argrelextrema to compute the local minima and maxima points
    local_min = argrelextrema(data.iloc[:-argrel_window]['Low'].values, np.less, order=argrel_window)[0]
    local_max = argrelextrema(data.iloc[:-argrel_window]['High'].values, np.greater, order=argrel_window)[0]

    # Store the minima and maxima values in a dataframe
    minima = data.iloc[local_min].Low
    maxima = data.iloc[local_max].High

    # Return dataframes containing minima and maxima values
    return minima, maxima



# 1 & 2) Rising Wedge Entry Long Trade
def Rising_Wedge_Long(min_max, minima_prices, maxima_prices, frequency='daily'):
    patterns = []

    for i in range(4, len(min_max)):
        window = min_max.iloc[i-4:i]

        if frequency == 'daily':
            window_size = (window.index[-1] - window.index[0]).days
        else:
            window_size = (window.index[-1] - window.index[0]).seconds / 60 / 60

        if window_size > 100:
            continue

        a, b, c, d, = [window[i] for i in range(0, len(window))]
        
        cond_1 = a in maxima_prices.values and c in maxima_prices.values 
        cond_2 = a < c
        cond_3 = b in minima_prices.values and d in minima_prices.values 
        cond_4 = b < d 
        #cond_5 = abs(a - c) <= np.mean([a, c]) * 0.015
        #cond_6 = abs(b - d) <= np.mean([b, d]) * 0.015
        
        if cond_1 and cond_2 and cond_3 and cond_4:
            patterns.append(([window.index[i] for i in range(0, len(window))]))
            
    return patterns

# 3) Rising Wedge 3 Point long Entry scanner function
def Rising_wedge_3_point_long_entry(min_max, minima_prices, maxima_prices, frequency='daily'):

    # To store pattern instances
    patterns = []

    # Loop to iterate along the price data
    for i in range(3, len(min_max)):

        # Store 3 local minima and local maxima points at a time in the variable 'window'
        window = min_max.iloc[i-3:i]

        # Determine window length based on the frequency of data
        if frequency == 'daily':
            window_size = (window.index[-1] - window.index[0]).days
        else:
            window_size = (window.index[-1] - window.index[0]).seconds/60/60

        # Ensure that pattern is formed within 100 bars
        if window_size > 100:
            continue

        # Store the 3 unique points to check for conditions
        a, b, c = [window[i] for i in range(0, len(window))]

        # cond_1: To check b is in maxima_prices
        cond_1 = b in maxima_prices.values

        # cond_2: To check a,c are in minima_prices
        cond_2 = all(x in minima_prices.values for x in [a, c])

        # cond_3: To check if the bottoms are below the neckline
        cond_3 = (a < c)       

        # Checking if all conditions are true
        if cond_1 and cond_2 and cond_3:

            # Append the pattern to list if all conditions are met
            patterns.append(([window.index[i] for i in range(0, len(window))]))

    return patterns


# Scanning All the Patterns
def scanning_patterns():
    
    #global data_list 
  

    # For Long Trades
    global RISING_WEDGE_A_LONG_result_dict        
    global RISING_WEDGE_C_LONG_result_dict     
    global RISING_WEDGE_3P_LONG_result_dict   
    
       
       
    # Clear previous data    
    #data_list.clear()
    
    # For Long Trades
    RISING_WEDGE_A_LONG_result_dict.clear()      
    RISING_WEDGE_C_LONG_result_dict.clear()       
    RISING_WEDGE_3P_LONG_result_dict.clear()     
    
     
    
    for ticker in Tickers:
        
        retry_count = 0
        success = False

        while retry_count < 3 and not success:
        
            try:                 
                                   
                data = data_list[ticker]      
                data.index = pd.to_datetime(data.index,format ="%Y-%m-%d %H:%M:%S")
                print('Data is Downloaded for ',ticker)
                
                # Long Trades
                long_new_minima, long_new_maxima = get_min_max(data, long_argrel_window)
                LONG_min_max = pd.concat([long_new_minima, long_new_maxima]).sort_index()
                
                
                # Check if the ticker is in the respective pattern ticker list before scanning
                if ticker in RISING_WEDGE_A_LONG_Tickers:               
                
                    try:
                        print("Scanning for Rising Wedge Entry For ",ticker)
                        Rising_wedge_pattern = Rising_Wedge_Long(LONG_min_max, long_new_minima, long_new_maxima)
                        Rising_wedge_scanner_data = pd.DataFrame(Rising_wedge_pattern)
                        Rising_wedge_scanner_data.columns = ['a', 'b', 'c', 'd']
                        Rising_wedge_scanner_data['a_price'] = data.loc[Rising_wedge_scanner_data.a, 'High'].values
                        Rising_wedge_scanner_data['b_price'] = data.loc[Rising_wedge_scanner_data.b, 'Low'].values
                        Rising_wedge_scanner_data['c_price'] = data.loc[Rising_wedge_scanner_data.c, 'High'].values
                        Rising_wedge_scanner_data['d_price'] = data.loc[Rising_wedge_scanner_data.d, 'Low'].values
                        RISING_WEDGE_A_LONG_result_dict[ticker] = Rising_wedge_scanner_data
                    except ValueError:
                        print(f"No Rising Wedge Entry Pattern Found in {ticker}")
                    
                       
                            
                 
                if ticker in RISING_WEDGE_C_LONG_Tickers:  
                    
                    try:
                        print("Scanning for Rising Wedge Entry For ",ticker)
                        Rising_wedge_pattern = Rising_Wedge_Long(LONG_min_max, long_new_minima, long_new_maxima)
                        Rising_wedge_scanner_data = pd.DataFrame(Rising_wedge_pattern)
                        Rising_wedge_scanner_data.columns = ['a', 'b', 'c', 'd']
                        Rising_wedge_scanner_data['a_price'] = data.loc[Rising_wedge_scanner_data.a, 'High'].values
                        Rising_wedge_scanner_data['b_price'] = data.loc[Rising_wedge_scanner_data.b, 'Low'].values
                        Rising_wedge_scanner_data['c_price'] = data.loc[Rising_wedge_scanner_data.c, 'High'].values
                        Rising_wedge_scanner_data['d_price'] = data.loc[Rising_wedge_scanner_data.d, 'Low'].values
                        RISING_WEDGE_C_LONG_result_dict[ticker] = Rising_wedge_scanner_data
                    except ValueError:
                        print(f"No Rising Wedge Entry Pattern Found in {ticker}")
                    
                                    
                        
                if ticker in RISING_WEDGE_3P_LONG_Tickers:  
                    
                    try:
                        print("Scanning for Rising Wedge 3 point Entry For ",ticker)
                        
                        Rising_wedge_3_point_pattern = Rising_wedge_3_point_long_entry(LONG_min_max, long_new_minima, long_new_maxima)  
                        
                        # Create a DataFrame for the patterns for each ticker
                        Rising_wedge_3_point_scanner_data = pd.DataFrame(Rising_wedge_3_point_pattern)
                        
                        # Rename columns for the 7 points in the pattern
                        Rising_wedge_3_point_scanner_data.columns = ['a', 'b', 'c']
                        
                        # Add price columns for each point in the pattern
                        Rising_wedge_3_point_scanner_data['a_price'] = data.loc[Rising_wedge_3_point_scanner_data.a, 'Low'].values
                        Rising_wedge_3_point_scanner_data['b_price'] = data.loc[Rising_wedge_3_point_scanner_data.b, 'High'].values
                        Rising_wedge_3_point_scanner_data['c_price'] = data.loc[Rising_wedge_3_point_scanner_data.c, 'Low'].values
                        RISING_WEDGE_3P_LONG_result_dict[ticker] = Rising_wedge_3_point_scanner_data
                    except ValueError:
                        print(f"No Rising Wedge 3 point Long Entry Pattern Found in {ticker}")                 
                
                success = True  #  Break retry loop                        
                        
            except Exception as e:
                print(f"Error processing {ticker}: {e}")
                retry_count += 1
                print(f"Error processing {ticker} (attempt {retry_count}/3): {e}")
                if retry_count >= 3:
                    print(f"Failed to process {ticker} after 3 attempts.")            
                continue

# 1) Rising Wedge Continuation A Entry Trading Parameter
Rising_Wedge_Continuation_A_Entry_tradingparameter = 'Rising_Wedge_Continuation_A_Long_Entry_Trading_parameters.xlsx'
Rising_Wedge_Continuation_A_Entry_tradefile = 'Rising_Wedge_Continuation_A_Entry_cash_Long_Trades.xlsx'

# 2) Rising Wedge Continuation C Entry Trading Parameter
Rising_Wedge_Continuation_C_Entry_tradingparameter = 'Rising_Wedge_Continuation_C_Long_Entry_Trading_parameters.xlsx'
Rising_Wedge_Continuation_C_Entry_tradefile = 'Rising_Wedge_Continuation_C_Entry_cash_Long_Trades.xlsx'

# 3) Rising Wedge 3 Buy Point Trading Parameter
Rising_Wedge_3p_Entry_tradingparameter = 'Rising_Wedge_3p_Long_Entry_Trading_parameters.xlsx'
Rising_Wedge_3p_Entry_tradefile = 'Rising_Wedge_3p_cash_Long_Trades.xlsx'


# Merge all long Pattern Option Selling
def all_long_trades_file():
    try:
        # Define the specific filenames you want to merge
        file_paths = [
            Rising_Wedge_Continuation_A_Entry_tradefile,
            Rising_Wedge_Continuation_C_Entry_tradefile,
            Rising_Wedge_3p_Entry_tradefile            
        ]

        # Load all the Excel files into a list of dataframes
        dfs = [pd.read_excel(file_path) for file_path in file_paths]

        # Merge all dataframes into one
        merged_df = pd.concat(dfs, ignore_index=True)

        # Ensure 'Entry Time' column is in datetime format
        merged_df['Entry Time'] = pd.to_datetime(merged_df['Entry Time'], errors='coerce')

        # Sort the merged dataframe based on Entry Time
        merged_df = merged_df.sort_values(by='Entry Time')
        
        merged_df = merged_df[(merged_df['Trade Status'] == 'OPEN')]
        
        current_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
        

        # Save the merged and sorted dataframe to a new Excel file
        output_file_path = 'All_Long_Open_Trades.xlsx'
        #merged_df.to_excel(output_file_path, index=False)

        print(f"Merged and sorted trades saved to {output_file_path} Time: {current_time}")
    except Exception as e:
        print(f"Error in All_Long_Open_Trades: {e}")
        
    return merged_df

# Create Excel Sheet only of it Needed or first time
def create_excel_sheet(filename):
    if not os.path.exists(filename):
        columns = ['ID', 'Symbol', 'Entry Time','Buy Price', 'Qty','Target Price','Exit Time', 'Sell Price','Points','Brokerage','Profit/Loss', 'Trade Status','Duration(Days)']
        df = pd.DataFrame(columns=columns)
        df.to_excel(filename, index=False)
        print(f"{filename} created successfully!")
    else:
        print(f"{filename} already exists.") 

# For Future Trade Files
create_excel_sheet(Rising_Wedge_Continuation_A_Entry_tradefile) # 1

create_excel_sheet(Rising_Wedge_Continuation_C_Entry_tradefile) # 2

create_excel_sheet(Rising_Wedge_3p_Entry_tradefile) # 3


# 1) Rising Wedge Continuation A Entry
def Rising_Wedge_Continuation_A_Entry_save_parameters_to_excel(result_dict, file_name):
    parameters_data = []
    
    # Generate parameters data
    for ticker, results in result_dict.items():
        for index, row in results.iterrows():
            candle_a = str(row['a'])
            candle_b = str(row['b'])
            candle_c = str(row['c'])
            candle_d = str(row['d'])
            
            a_price = row['a_price']
            b_price = row['b_price']
            c_price = row['c_price'] 
            d_price = row['d_price'] 
            
            
            
            start_date = str(row['a'])
            end_date = str(row['d'])
            entry_price = row['a_price']
            points = min((row['a_price'] - row['b_price']),(row['c_price'] - row['d_price']))
            stop_loss_price = row['b_price']
            
            parameters_data.append({
                'ID': 'Rising_Wedge_A_LONG'+ticker + start_date,
                'Symbol': ticker,
                'a':candle_a,
                'b':candle_b,
                'c':candle_c,
                'd':candle_d,
                     
                'a_price':a_price,
                'b_price':b_price,
                'c_price':c_price,
                'd_price':d_price,
                
                'Start_Date': start_date,
                'End_Date': end_date,
                'Points': points,
                'Entry_Price': entry_price,
                'SL_Price': stop_loss_price
            })
    
    # Create a DataFrame from the parameters_data list
    new_data_df = pd.DataFrame(parameters_data)

    # Add new columns 'Trade Status' and 'Action' with no values
    new_data_df['Trade_Status'] = np.nan
    new_data_df['Action'] = np.nan

    # Read existing data from Excel file
    try:
        existing_data_df = pd.read_excel(file_name)
    except FileNotFoundError:
        existing_data_df = pd.DataFrame()

    # Merge new data with existing data and remove duplicates
    updated_data_df = pd.concat([existing_data_df, new_data_df]).drop_duplicates(subset='ID', keep='first')

    # Save the updated DataFrame to Excel
    updated_data_df.to_excel(file_name, index=False)

# 2) Rising Wedge Continuation C Entry Saving Trading parameters
def Rising_Wedge_Continuation_C_Entry_save_parameters_to_excel(result_dict, file_name):
    parameters_data = []
    
    # Generate parameters data
    for ticker, results in result_dict.items():
        for index, row in results.iterrows():
            candle_a = str(row['a'])
            candle_b = str(row['b'])
            candle_c = str(row['c'])
            candle_d = str(row['d'])
            
            a_price = row['a_price']
            b_price = row['b_price']
            c_price = row['c_price'] 
            d_price = row['d_price'] 
            
            
            
            start_date = str(row['a'])
            end_date = str(row['d'])
            entry_price = row['c_price']
            points = min((row['a_price'] - row['b_price']),(row['c_price'] - row['d_price']))
            stop_loss_price = row['b_price']
            
            parameters_data.append({
                'ID': 'Rising_Wedge_C_LONG'+ticker + start_date,
                'Symbol': ticker,
                'a':candle_a,
                'b':candle_b,
                'c':candle_c,
                'd':candle_d,
                     
                'a_price':a_price,
                'b_price':b_price,
                'c_price':c_price,
                'd_price':d_price,
                
                'Start_Date': start_date,
                'End_Date': end_date,
                'Points': points,
                'Entry_Price': entry_price,
                'SL_Price': stop_loss_price
            })
    
    # Create a DataFrame from the parameters_data list
    new_data_df = pd.DataFrame(parameters_data)

    # Add new columns 'Trade Status' and 'Action' with no values
    new_data_df['Trade_Status'] = np.nan
    new_data_df['Action'] = np.nan

    # Read existing data from Excel file
    try:
        existing_data_df = pd.read_excel(file_name)
    except FileNotFoundError:
        existing_data_df = pd.DataFrame()

    # Merge new data with existing data and remove duplicates
    updated_data_df = pd.concat([existing_data_df, new_data_df]).drop_duplicates(subset='ID', keep='first')

    # Save the updated DataFrame to Excel
    updated_data_df.to_excel(file_name, index=False)

# 3) Rising Wedge 3p Entry Saving Trading parameters    
def Rising_Wedge_3p_Long_Entry_save_parameters_to_excel(result_dict, file_name):
    parameters_data = []
    
    # Generate parameters data
    for ticker, results in result_dict.items():
        for index, row in results.iterrows():
            candle_a = str(row['a'])
            candle_b = str(row['b'])
            candle_c = str(row['c'])
            a_price = row['a_price']
            b_price = row['b_price']
            c_price = row['c_price']            
            start_date = str(row['a'])
            end_date = str(row['c'])
            entry_price = row['b_price']
            points = row['b_price'] - min(row['a_price'],row['c_price'])
            stop_loss_price = min(row['a_price'],row['c_price'])
            
            parameters_data.append({
                'ID': 'Rising_Wedge_3p_Entry_LONG'+ticker + start_date,
                'Symbol': ticker,
                'a':candle_a,
                'b':candle_b,
                'c':candle_c,
                'a_price':a_price,
                'b_price':b_price,
                'c_price':c_price,
                'Start_Date': start_date,
                'End_Date': end_date,
                'Points': points,
                'Entry_Price': entry_price,
                'SL_Price': stop_loss_price
            })
    
    # Create a DataFrame from the parameters_data list
    new_data_df = pd.DataFrame(parameters_data)

    # Add new columns 'Trade Status' and 'Action' with no values
    new_data_df['Trade_Status'] = np.nan
    new_data_df['Action'] = np.nan

    # Read existing data from Excel file
    try:
        existing_data_df = pd.read_excel(file_name)
    except FileNotFoundError:
        existing_data_df = pd.DataFrame()

    # Merge new data with existing data and remove duplicates
    updated_data_df = pd.concat([existing_data_df, new_data_df]).drop_duplicates(subset='ID', keep='first')

    # Save the updated DataFrame to Excel
    updated_data_df.to_excel(file_name, index=False)
 

# Update pattern Status whether it is open for trade or Time Out    
def update_Long_pattern_status(file_name, data_dict):
    try:
        df = pd.read_excel(file_name)
    except FileNotFoundError:
        print("File not found")
        return

    # Update 'Trade_Status' column based on conditions for rows with NaN values or 'OPEN'
    for index, row in df.iterrows():
        if pd.isna(row['Trade_Status']):
            symbol = row['Symbol']
            end_date = row['End_Date']
            entry_price = row['Entry_Price']

            if symbol in data_dict:
                data = data_dict[symbol]
                try:
                    last_candle_date = data.index[-1]
                    if last_candle_date > pd.to_datetime(end_date):
                        last_candle_index = data.index.get_loc(pd.to_datetime(end_date)) + 1
                        if (data.loc[pd.to_datetime(end_date):last_candle_date, 'High'] >= entry_price).any():
                            df.loc[index, 'Trade_Status'] = 'TIME OUT'
                        else:
                            df.loc[index, 'Trade_Status'] = 'OPEN'
                    else:
                        print(f"Data not available for the period from {end_date} to {last_candle_date} for symbol {symbol}.")
                except KeyError:
                    print(f"Date {end_date} not found in the data for symbol {symbol}.")

        elif row['Trade_Status'] == 'OPEN':
            symbol = row['Symbol']
            end_date = row['End_Date']
            entry_price = row['Entry_Price']

            if symbol in data_dict:
                data = data_dict[symbol]
                try:
                    last_candle_date = data.index[-1]
                    remaining_candles = len(data.loc[pd.to_datetime(end_date):])
                    if remaining_candles <= 177:
                        df.loc[index, 'Trade_Status'] = 'OPEN'
                    else:
                        df.loc[index, 'Trade_Status'] = 'TIME OUT'
                except KeyError:
                    print(f"Date {end_date} not found in the data for symbol {symbol}.")

    df.to_excel(file_name, index=False)




# Update Trade Status Trade Taken
def update_trade_status(filename, trade_id):
    workbook = load_workbook(filename=filename)
    worksheet = workbook.active

    # Find the index of the 'Action' column
    action_col_idx = None
    for cell in worksheet[1]:  # assuming the first row contains headers
        if cell.value == 'Action':
            action_col_idx = cell.column
            break

    if action_col_idx is None:
        raise ValueError('Action column not found')

    for row in worksheet.iter_rows():
        if row[0].value == trade_id:
            row[action_col_idx - 1].value = 'Trade Taken'  # subtract 1 because row is 0-indexed
            break

    workbook.save(filename)  
    
# Update Long trade entry
def update_long_trades(trade_id, ticker,entry_time,cash_buy,quantity,target_price,tradefile):
    workbook = load_workbook(filename=tradefile)
    worksheet = workbook.active

    new_row = [trade_id, ticker,entry_time,cash_buy,quantity, target_price,'','','','','', 'OPEN','']
    worksheet.append(new_row)

    workbook.save(tradefile)
    
   


# Update Long Trade Target Hit
def update_long_trades_target_exit(trade_id,exit_time,buy_exit, points,brokerage, profit_loss,tradefile):
    # Load the Long_Trades.xlsx file
    wb = load_workbook(tradefile)
    ws = wb.active
    
    # Initialize trade_row
    trade_row = None
    entry_time = None
    
    # Find the row with the given trade_id
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            if cell.value == trade_id:
                trade_row = cell.row
                entry_time = ws.cell(row=trade_row, column=3).value  # Assuming Entry Time is in column 5
                break
    
    # Check if trade_row is assigned a value
    if trade_row is not None and entry_time is not None:
        # Update the values in the row
        ws.cell(row=trade_row, column=7).value = exit_time
        ws.cell(row=trade_row, column=8).value = buy_exit
        ws.cell(row=trade_row, column=9).value = points
        ws.cell(row=trade_row, column=10).value = brokerage
        ws.cell(row=trade_row, column=11).value = profit_loss
        ws.cell(row=trade_row, column=12).value = 'Target Hit'
        
        # Calculate duration in days
        entry_time = dt.datetime.strptime(entry_time, "%d-%b-%Y %I:%M%p")
        exit_time = dt.datetime.strptime(exit_time, "%d-%b-%Y %I:%M%p")
        duration_days = (exit_time - entry_time).days
        ws.cell(row=trade_row, column=13).value = duration_days  # Assuming Duration(Days) is in column 17

        # Save the updated workbook
        wb.save(tradefile)
    else:
        print("Trade ID not found")    


    
# Update Pattern Status after closing the trade 
def update_trade_status_with_trade_closed(file_path, trade_id,col):
    # Load the trading_parameters.xlsx file
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find the row with the given trade_id
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value == trade_id:
                trade_row = cell.row
                break
    
    # Update the Action column value in the row
    ws.cell(row=trade_row, column=col).value = 'Trade Closed'  # Replace 3 with the correct column number for "Action"
    
    # Save the updated workbook
    wb.save(file_path)
    
    
def Cash_Trade_Long_Only():
    
    scanning_patterns()

    Rising_Wedge_Continuation_A_Entry_save_parameters_to_excel(RISING_WEDGE_A_LONG_result_dict, Rising_Wedge_Continuation_A_Entry_tradingparameter)

    Rising_Wedge_Continuation_C_Entry_save_parameters_to_excel(RISING_WEDGE_C_LONG_result_dict, Rising_Wedge_Continuation_C_Entry_tradingparameter)

    Rising_Wedge_3p_Long_Entry_save_parameters_to_excel(RISING_WEDGE_3P_LONG_result_dict, Rising_Wedge_3p_Entry_tradingparameter)



    # Long Trades
    # 1) Rising Wedge Continuation A Long Entry
    update_Long_pattern_status(Rising_Wedge_Continuation_A_Entry_tradingparameter, data_list)
    update_Long_pattern_status(Rising_Wedge_Continuation_A_Entry_tradingparameter, data_list)

    # 2) Rising Wedge Continuation C Long Entry
    update_Long_pattern_status(Rising_Wedge_Continuation_C_Entry_tradingparameter, data_list)
    update_Long_pattern_status(Rising_Wedge_Continuation_C_Entry_tradingparameter, data_list)


    # 3) Rising Wedge 3 Point Long Parameter
    update_Long_pattern_status(Rising_Wedge_3p_Entry_tradingparameter, data_list)
    update_Long_pattern_status(Rising_Wedge_3p_Entry_tradingparameter, data_list)




    # Long trades
    # 1) For Rising Wedge Continuation A Entry
    Rising_Wedge_Continuation_A_Entry_df = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradingparameter)
    # Reading the Long_trade Details
    Rising_Wedge_Continuation_A_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradefile)

    # 2) For Rising Wedge Continuation C Entry
    Rising_Wedge_Continuation_C_Entry_df = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradingparameter)
    # Reading the Long_trade Details
    Rising_Wedge_Continuation_C_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradefile)


    # 3) For Rising Wedge 3 Point
    Rising_Wedge_3p_Entry_df = pd.read_excel(Rising_Wedge_3p_Entry_tradingparameter)
    # Reading the Long_trade Details
    Rising_Wedge_3p_Entry_dft = pd.read_excel(Rising_Wedge_3p_Entry_tradefile)

     
    endTime = dt.datetime.now(pytz.timezone('Asia/Kolkata')).replace(hour=EXIT_TIME[0], minute=EXIT_TIME[1],second=EXIT_TIME[2])
    while dt.datetime.now(pytz.timezone('Asia/Kolkata')) <  endTime:

        try:

            for i in Tickers:

               #spot_prices[i] = spot(instrument_lookup(instrument=instrument_df, symbol=i))


               # print("#######################################################")
               # print("###########  CASH STOCK TRADES WITH HEDGES ############")
               # print("#######################################################")



               # print('Spot prices of',i,' ',spot_prices[i])

               #time.sleep(0.25)

               # Updating Candles only on a specific time
               if is_required_time():
                   
                   time.sleep(75)

                   scanning_patterns()
                   
                   Rising_Wedge_Continuation_A_Entry_save_parameters_to_excel(RISING_WEDGE_A_LONG_result_dict, Rising_Wedge_Continuation_A_Entry_tradingparameter)

                   Rising_Wedge_Continuation_C_Entry_save_parameters_to_excel(RISING_WEDGE_C_LONG_result_dict, Rising_Wedge_Continuation_C_Entry_tradingparameter)

                   Rising_Wedge_3p_Long_Entry_save_parameters_to_excel(RISING_WEDGE_3P_LONG_result_dict, Rising_Wedge_3p_Entry_tradingparameter)

                   

                   # Long Trades
                   # 1) Rising Wedge Continuation A Long Entry
                   update_Long_pattern_status(Rising_Wedge_Continuation_A_Entry_tradingparameter, data_list)
                   update_Long_pattern_status(Rising_Wedge_Continuation_A_Entry_tradingparameter, data_list)

                   # 2) Rising Wedge Continuation C Long Entry
                   update_Long_pattern_status(Rising_Wedge_Continuation_C_Entry_tradingparameter, data_list)
                   update_Long_pattern_status(Rising_Wedge_Continuation_C_Entry_tradingparameter, data_list)


                   # 3) Rising Wedge 3 Point Long Parameter
                   update_Long_pattern_status(Rising_Wedge_3p_Entry_tradingparameter, data_list)
                   update_Long_pattern_status(Rising_Wedge_3p_Entry_tradingparameter, data_list)


                   
                   # Long trades
                   # 1) For Rising Wedge Continuation A Entry
                   Rising_Wedge_Continuation_A_Entry_df = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradingparameter)
                   # Reading the Long_trade Details
                   Rising_Wedge_Continuation_A_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradefile)

                   # 2) For Rising Wedge Continuation C Entry
                   Rising_Wedge_Continuation_C_Entry_df = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradingparameter)
                   # Reading the Long_trade Details
                   Rising_Wedge_Continuation_C_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradefile)


                   # 3) For Rising Wedge 3 Point
                   Rising_Wedge_3p_Entry_df = pd.read_excel(Rising_Wedge_3p_Entry_tradingparameter)
                   # Reading the Long_trade Details
                   Rising_Wedge_3p_Entry_dft = pd.read_excel(Rising_Wedge_3p_Entry_tradefile)

                   
                   print("#######################")
                   print("### Candle Updation ###")
                   print("#######################")

                   time.sleep(20)
                   
                   
                   
               # 1) For Rising Wedge Continuation A Entry
               Rising_Wedge_Continuation_A_Entry_valid_trades = Rising_Wedge_Continuation_A_Entry_df[(Rising_Wedge_Continuation_A_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_Continuation_A_Entry_df['Action']))]

               # 1) For Long Reversal A Entry
               Rising_Wedge_Continuation_A_Entry_trade_row = Rising_Wedge_Continuation_A_Entry_valid_trades[Rising_Wedge_Continuation_A_Entry_valid_trades['Symbol'] == i]
               Rising_Wedge_Continuation_A_Entry_trade_row = Rising_Wedge_Continuation_A_Entry_trade_row.sort_values(by="Entry_Price", ascending=True)


               # 2) For Rising Wedge Continuation C Entry
               Rising_Wedge_Continuation_C_Entry_valid_trades = Rising_Wedge_Continuation_C_Entry_df[(Rising_Wedge_Continuation_C_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_Continuation_C_Entry_df['Action']))]
               # 2) For Rising Wedge Continuation C Entry
               Rising_Wedge_Continuation_C_Entry_trade_row = Rising_Wedge_Continuation_C_Entry_valid_trades[Rising_Wedge_Continuation_C_Entry_valid_trades['Symbol'] == i]
               Rising_Wedge_Continuation_C_Entry_trade_row = Rising_Wedge_Continuation_C_Entry_trade_row.sort_values(by="Entry_Price", ascending=True)


               # 3) For Rising Wedge 3 Point Buy
               Rising_Wedge_3p_Entry_valid_trades = Rising_Wedge_3p_Entry_df[(Rising_Wedge_3p_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_3p_Entry_df['Action']))]
               # 3) For Rising Wedge 3 Point
               Rising_Wedge_3p_Entry_trade_row = Rising_Wedge_3p_Entry_valid_trades[Rising_Wedge_3p_Entry_valid_trades['Symbol'] == i]
               Rising_Wedge_3p_Entry_trade_row = Rising_Wedge_3p_Entry_trade_row.sort_values(by="Entry_Price", ascending=True)

               
               
               ################################################################
               ########## Rising Wedge Continuation A Entry ###################
               ################################################################                

               # Checking For Long Entry

               if not Rising_Wedge_Continuation_A_Entry_trade_row.empty:
                   entry_point = Rising_Wedge_Continuation_A_Entry_trade_row['Entry_Price'].iloc[0]

                   # Step 3: Compare spot price with entry point
                   if spot_prices[i] > entry_point:
                       
                       five_min_close_price = min_30_close(i)
                         
                                                  
                       if five_min_close_price > entry_point:
                           
                                           
                           
                            
                           open_trades_df = all_long_trades_file()
                            
                           filtered_df = open_trades_df[open_trades_df['Trade Status'] == 'OPEN']
                            
                           open_trade_count = len(filtered_df)
                        
                           # Check if not more than give open trades
                         
                           if open_trade_count >= Long_max_open_trade:
                               print(f"Skipping trade for {Rising_Wedge_Continuation_A_Entry_trade_row['ID'].iloc[0]} as open trade count is {open_trade_count}.")
                               tele_msg(f"Skipping trade for {Rising_Wedge_Continuation_A_Entry_trade_row['ID'].iloc[0]} as open trade count is {open_trade_count}.")
                         
                               # Updating Parameter Excel
                               trade_id = Rising_Wedge_Continuation_A_Entry_trade_row['ID'].iloc[0]
                         
                               Rising_Wedge_Continuation_A_Entry_df.loc[Rising_Wedge_Continuation_A_Entry_df['ID'] == trade_id, 'Trade_Status'] = 'TIME OUT'
                         
                               Rising_Wedge_Continuation_A_Entry_df.to_excel(Rising_Wedge_Continuation_A_Entry_tradingparameter, index=False)
                                                
                            
                           else:
                                # Send buy order to API
                                trade_id = Rising_Wedge_Continuation_A_Entry_trade_row['ID'].iloc[0]
                                # Replace this line with your API call to send a buy order
                                 
                                current_price = float(spot_prices[i])
                                                        
                                qty = int(math.floor(Total_Cash_per_position/current_price))
                                 
                                entry_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                
                                # ZERODHA PLACING BUY ORDER AND GET EXECUTED PRICE FROM POSITION BOOK
                                
                                ###
                                
                                kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= i, transaction_type='BUY', quantity=qty, product='CNC', order_type='MARKET')
                                
                                time.sleep(15)
                                
                                position_df = pd.DataFrame(kite.positions()["net"])
                                
                                Long_avg_price = float(position_df[position_df.tradingsymbol==i].buy_price.values[0]) 
                                
                                ###
                                
                                                                                            
                                future_entry = float(Long_avg_price) #float(position_df[position_df.tradingsymbol==fut].buy_price.values[0])                                           
                                                        
                                
                                # Step 4: Update Long_Trades.xlsx
                                point = Rising_Wedge_Continuation_A_Entry_trade_row['Points'].iloc[0]
                                 
                                target_price = int(math.ceil(future_entry + (point * RISING_WEDGE_A_LONG_targetx)))
                                 
                                 
                                 
                                update_long_trades(trade_id, i,entry_time,Long_avg_price,qty,target_price,Rising_Wedge_Continuation_A_Entry_tradefile)
                         
                                # Update trading_parameters.xlsx
                                update_trade_status(Rising_Wedge_Continuation_A_Entry_tradingparameter, trade_id)
                                 
                                time.sleep(3)
                                 
                                Rising_Wedge_Continuation_A_Entry_df = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradingparameter)  
                                 
                                Rising_Wedge_Continuation_A_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradefile)
                                 
                                time.sleep(3)
                                 
                                # Step 1 check Trade_Status = 'OPEN' and Action = ''     
                                Rising_Wedge_Continuation_A_Entry_valid_trades = Rising_Wedge_Continuation_A_Entry_df[(Rising_Wedge_Continuation_A_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_Continuation_A_Entry_df['Action']))]
                                                        
                                # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                                Rising_Wedge_Continuation_A_Entry_open_trades = Rising_Wedge_Continuation_A_Entry_dft[(Rising_Wedge_Continuation_A_Entry_dft['Trade Status'] == 'OPEN')]
                                                        
                                # Check Action column = 'Trade Taken'
                                Rising_Wedge_Continuation_A_Entry_taken_trades = Rising_Wedge_Continuation_A_Entry_df[Rising_Wedge_Continuation_A_Entry_df['Action'] == 'Trade Taken']
                                 
                                tele_msg(' Rising Wedge Continuation A Entry '+i+' '+' Long entry avg price '+str(Long_avg_price)+' Target '+str(target_price)+' And Qty '+ str(qty))
                                 
                                   
               
                                        
                                           
               # Checking Target Hit
               # Check Action column = 'Trade Taken'
               Rising_Wedge_Continuation_A_Entry_taken_trades = Rising_Wedge_Continuation_A_Entry_df[Rising_Wedge_Continuation_A_Entry_df['Action'] == 'Trade Taken']

               # Loop through each taken trade
               for index, row in Rising_Wedge_Continuation_A_Entry_taken_trades.iterrows():
                   trade_id = row['ID']
                   
                   # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                   Rising_Wedge_Continuation_A_Entry_trade = Rising_Wedge_Continuation_A_Entry_dft[(Rising_Wedge_Continuation_A_Entry_dft['ID'] == trade_id) & (Rising_Wedge_Continuation_A_Entry_dft['Trade Status'] == 'OPEN')]
                   
                   if not Rising_Wedge_Continuation_A_Entry_trade.empty:
                       sym = Rising_Wedge_Continuation_A_Entry_trade['Symbol'].iloc[0]
                       if sym == i:  # Check if the ticker matches the symbol from the Excel sheet
                       
                           target_price = Rising_Wedge_Continuation_A_Entry_trade['Target Price'].iloc[0]
                
                           # Compare spot price with stop loss price
                           if spot_prices[i] > target_price:
                               
                               qty = int(Rising_Wedge_Continuation_A_Entry_trade['Qty'].iloc[0])
                               
                               fut = Rising_Wedge_Continuation_A_Entry_trade['Symbol'].iloc[0]
                
                               # Send sell order to API
                               # Replace this line with your API call to send a sell order
                               
                               # Send sell order to API
                               # Replace this line with your API call to send a sell order                              

                               exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")  
                               
                               
                               # ZERODHA LONG TRADE EXIT AND BOOKING PROFITS
                               
                               ###
                               
                               kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= sym, transaction_type='SELL', quantity=qty, product='CNC', order_type="LIMIT",price = target_price)
                                                    
                               time.sleep(15)
                                                         
                                   
                               position_df = pd.DataFrame(kite.positions()["net"])
                               
                               Long_exit_avg_price = float(position_df[position_df.tradingsymbol==i].sell_price.values[0])  
                               
                               ###
                                                                                         
                               # Update Long_Trades.xlsx
                               buy_exit = Long_exit_avg_price # float(position_df[position_df.tradingsymbol==fut].sell_price.values[0])                                           
                                 
                               buy_entry = float(Rising_Wedge_Continuation_A_Entry_trade['Buy Price'].iloc[0])
                               
                               points = buy_exit - buy_entry
                               
                               broker = (((qty*buy_entry)+(qty*buy_exit)) * 0.0011)
                               
                               profit_loss = ((points * qty) - broker)
                
                               update_long_trades_target_exit(trade_id,exit_time, buy_exit, points,broker,profit_loss,Rising_Wedge_Continuation_A_Entry_tradefile)
                
                               update_trade_status_with_trade_closed(Rising_Wedge_Continuation_A_Entry_tradingparameter, trade_id,17)
                               
                               time.sleep(3)
                               
                               Rising_Wedge_Continuation_A_Entry_df = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradingparameter)     
                               
                               Rising_Wedge_Continuation_A_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_A_Entry_tradefile)
                               
                               time.sleep(3)
                               
                               # Step 1 check Trade_Status = 'OPEN' and Action = ''     
                               Rising_Wedge_Continuation_A_Entry_valid_trades = Rising_Wedge_Continuation_A_Entry_df[(Rising_Wedge_Continuation_A_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_Continuation_A_Entry_df['Action']))]
                               
                               
                               # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                               Rising_Wedge_Continuation_A_Entry_open_trades = Rising_Wedge_Continuation_A_Entry_dft[(Rising_Wedge_Continuation_A_Entry_dft['Trade Status'] == 'OPEN')]
                               
                               
                               # Check Action column = 'Trade Taken'
                               Rising_Wedge_Continuation_A_Entry_taken_trades = Rising_Wedge_Continuation_A_Entry_df[Rising_Wedge_Continuation_A_Entry_df['Action'] == 'Trade Taken']  
                               
                               tele_msg(' Rising Wedge Continuation A Entry '+i+' Tp Hit '+ ' Sold at Price '+str(buy_exit)+' And Profit Rupees '+str(profit_loss))

                
               ################################################################
               ########## Rising Wedge Continuation C Entry ################### 
               ################################################################                

               # Checking For Long Entry

               if not Rising_Wedge_Continuation_C_Entry_trade_row.empty:
                   entry_point = Rising_Wedge_Continuation_C_Entry_trade_row['Entry_Price'].iloc[0]

                   # Step 3: Compare spot price with entry point
                   if spot_prices[i] > entry_point:
                       
                       five_min_close_price = min_30_close(i)
                        
                       if five_min_close_price > entry_point:
             
                    
                           
                            
                           open_trades_df = all_long_trades_file()
                           
                           filtered_df = open_trades_df[open_trades_df['Trade Status'] == 'OPEN']
                            
                           open_trade_count = len(filtered_df)
                            
                           # Check if not more than give open trades
         
                           if open_trade_count >= Long_max_open_trade:
                               print(f"Skipping trade for {Rising_Wedge_Continuation_C_Entry_trade_row['ID'].iloc[0]} as open trade count is {open_trade_count}.")
                               tele_msg(f"Skipping trade for {Rising_Wedge_Continuation_C_Entry_trade_row['ID'].iloc[0]} as open trade count is {open_trade_count}.")
         
                               # Updating Parameter Excel
                               trade_id = Rising_Wedge_Continuation_C_Entry_trade_row['ID'].iloc[0]
         
                               Rising_Wedge_Continuation_C_Entry_df.loc[Rising_Wedge_Continuation_C_Entry_df['ID'] == trade_id, 'Trade_Status'] = 'TIME OUT'
         
                               Rising_Wedge_Continuation_C_Entry_df.to_excel(Rising_Wedge_Continuation_C_Entry_tradingparameter, index=False)
                            
                           else:
                                
                                # Send buy order to API
                                trade_id = Rising_Wedge_Continuation_C_Entry_trade_row['ID'].iloc[0]
                                 
                                # Replace this line with your API call to send a buy order
                                 
                                current_price = float(spot_prices[i])
                                                        
                                qty = int(math.floor(Total_Cash_per_position/current_price))
                                 
                                entry_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                
                                
                                # ZERODHA PLACING BUY ORDER AND GET EXECUTED PRICE FROM POSITION BOOK
                                
                                ###
                                
                                kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= i, transaction_type='BUY', quantity=qty, product='CNC', order_type='MARKET')
                                
                                time.sleep(15)
                                
                                position_df = pd.DataFrame(kite.positions()["net"])
                                
                                Long_avg_price = float(position_df[position_df.tradingsymbol==i].buy_price.values[0]) 
                                
                                ###
                                
                                                                                                                          
                                future_entry = float(Long_avg_price) #float(position_df[position_df.tradingsymbol==fut].buy_price.values[0])                                             
                                                              
                                 
                                # Step 4: Update Long_Trades.xlsx
                                point = Rising_Wedge_Continuation_C_Entry_trade_row['Points'].iloc[0]
                                 
                                                           
                                target_price = int(math.ceil(future_entry + (point * RISING_WEDGE_C_LONG_targetx)))
                                 
                                
                                update_long_trades(trade_id, i,entry_time,Long_avg_price,qty,target_price,Rising_Wedge_Continuation_C_Entry_tradefile)
                                 
                                # Update trading_parameters.xlsx
                                update_trade_status(Rising_Wedge_Continuation_C_Entry_tradingparameter, trade_id)
                                 
                                time.sleep(3)
                                 
                                Rising_Wedge_Continuation_C_Entry_df = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradingparameter)  
                                 
                                Rising_Wedge_Continuation_C_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradefile)
                                 
                                time.sleep(3)
                                 
                                # Step 1 check Trade_Status = 'OPEN' and Action = ''     
                                Rising_Wedge_Continuation_C_Entry_valid_trades = Rising_Wedge_Continuation_C_Entry_df[(Rising_Wedge_Continuation_C_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_Continuation_C_Entry_df['Action']))]
                                                        
                                # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                                Rising_Wedge_Continuation_C_Entry_open_trades = Rising_Wedge_Continuation_C_Entry_dft[(Rising_Wedge_Continuation_C_Entry_dft['Trade Status'] == 'OPEN')]
                                                        
                                # Check Action column = 'Trade Taken'
                                Rising_Wedge_Continuation_C_Entry_taken_trades = Rising_Wedge_Continuation_C_Entry_df[Rising_Wedge_Continuation_C_Entry_df['Action'] == 'Trade Taken']
                                                            
                                tele_msg(' Rising Wedge Continuation C Entry '+i+' '+' Long entry avg price '+str(Long_avg_price)+' Target '+str(target_price)+' And Qty '+ str(qty))
                                 
                                    
                    
                                                     
               # Checking Target Hit
               # Check Action column = 'Trade Taken'
               Rising_Wedge_Continuation_C_Entry_taken_trades = Rising_Wedge_Continuation_C_Entry_df[Rising_Wedge_Continuation_C_Entry_df['Action'] == 'Trade Taken']

               # Loop through each taken trade
               for index, row in Rising_Wedge_Continuation_C_Entry_taken_trades.iterrows():
                   trade_id = row['ID']
                   
                   # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                   Rising_Wedge_Continuation_C_Entry_trade = Rising_Wedge_Continuation_C_Entry_dft[(Rising_Wedge_Continuation_C_Entry_dft['ID'] == trade_id) & (Rising_Wedge_Continuation_C_Entry_dft['Trade Status'] == 'OPEN')]
                   
                   if not Rising_Wedge_Continuation_C_Entry_trade.empty:
                       sym = Rising_Wedge_Continuation_C_Entry_trade['Symbol'].iloc[0]
                       if sym == i:  # Check if the ticker matches the symbol from the Excel sheet
                           target_price = Rising_Wedge_Continuation_C_Entry_trade['Target Price'].iloc[0]
                
                           # Compare spot price with stop loss price
                           if spot_prices[i] > target_price:
                               
                               qty = int(Rising_Wedge_Continuation_C_Entry_trade['Qty'].iloc[0])
                               
                               fut = Rising_Wedge_Continuation_C_Entry_trade['Symbol'].iloc[0]
                               
                                
                               # Send sell order to API
                               # Replace this line with your API call to send a sell order
                               
                               # Send sell order to API
                               # Replace this line with your API call to send a sell order                              

                               exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                               
                               # ZERODHA LONG TRADE EXIT AND BOOKING PROFITS
                               
                               ###
                               
                               kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= sym, transaction_type='SELL', quantity=qty, product='CNC', order_type="LIMIT",price = target_price)
                                                    
                               time.sleep(15)
                                                         
                                   
                               position_df = pd.DataFrame(kite.positions()["net"])
                               
                               Long_exit_avg_price = float(position_df[position_df.tradingsymbol==i].sell_price.values[0])  
                               
                               ###
                               
                                                          
                               # Update Long_Trades.xlsx
                               buy_exit = Long_exit_avg_price # float(position_df[position_df.tradingsymbol==fut].sell_price.values[0])                                           
                                 
                   
                               buy_entry = float(Rising_Wedge_Continuation_C_Entry_trade['Buy Price'].iloc[0])
                               
                               points = buy_exit - buy_entry
                               
                               broker = (((qty*buy_entry)+(qty*buy_exit)) * 0.0011)
                               
                               profit_loss = ((points * qty) - broker)
                
                               update_long_trades_target_exit(trade_id,exit_time, buy_exit, points,broker,profit_loss,Rising_Wedge_Continuation_C_Entry_tradefile)
                           
                               
                               update_trade_status_with_trade_closed(Rising_Wedge_Continuation_C_Entry_tradingparameter, trade_id,17)
                               
                               time.sleep(3)
                               
                               Rising_Wedge_Continuation_C_Entry_df = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradingparameter)     
                               
                               Rising_Wedge_Continuation_C_Entry_dft = pd.read_excel(Rising_Wedge_Continuation_C_Entry_tradefile)
                               
                               time.sleep(3)
                               
                               # Step 1 check Trade_Status = 'OPEN' and Action = ''     
                               Rising_Wedge_Continuation_C_Entry_valid_trades = Rising_Wedge_Continuation_C_Entry_df[(Rising_Wedge_Continuation_C_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_Continuation_C_Entry_df['Action']))]
                               
                               
                               # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                               Rising_Wedge_Continuation_C_Entry_open_trades = Rising_Wedge_Continuation_C_Entry_dft[(Rising_Wedge_Continuation_C_Entry_dft['Trade Status'] == 'OPEN')]
                               
                               
                               # Check Action column = 'Trade Taken'
                               Rising_Wedge_Continuation_C_Entry_taken_trades = Rising_Wedge_Continuation_C_Entry_df[Rising_Wedge_Continuation_C_Entry_df['Action'] == 'Trade Taken']  
                               
                               tele_msg(' Rising Wedge Continuation C Entry '+i+' Tp Hit '+ ' Sold at Price '+str(Long_exit_avg_price)+' And Profit Rupees '+str(profit_loss))

                                                              
                            
                              
               #######################################
               ######## Rising Wedge 3 Point Buy #####
               #######################################

               # Checking For Long Entry

               if not Rising_Wedge_3p_Entry_trade_row.empty:
                   entry_point = Rising_Wedge_3p_Entry_trade_row['Entry_Price'].iloc[0]

                   # Step 3: Compare spot price with entry point
                   if spot_prices[i] > entry_point:
                       
                       five_min_close_price = min_30_close(i)
                          
                           
                       if five_min_close_price > entry_point:                        
                                         
                    
                           
                           
                           open_trades_df = all_long_trades_file()
                           
                           filtered_df = open_trades_df[open_trades_df['Trade Status'] == 'OPEN']
                            
                           open_trade_count = len(filtered_df)
                    
                           # Check if not more than give open trades
                            
                           if open_trade_count >= Long_max_open_trade:
                               print(f"Skipping trade for {Rising_Wedge_3p_Entry_trade_row['ID'].iloc[0]} as open trade count is {open_trade_count}.")
                               tele_msg(f"Skipping trade for {Rising_Wedge_3p_Entry_trade_row['ID'].iloc[0]} as open trade count is {open_trade_count}.")
             
                               # Updating Parameter Excel
                               trade_id = Rising_Wedge_3p_Entry_trade_row['ID'].iloc[0]
             
                               Rising_Wedge_3p_Entry_df.loc[Rising_Wedge_3p_Entry_df['ID'] == trade_id, 'Trade_Status'] = 'TIME OUT'
             
                               Rising_Wedge_3p_Entry_df.to_excel(Rising_Wedge_3p_Entry_tradingparameter, index=False)
                            
                            
                           else:
                                
                                
                               # Send buy order to API
                               trade_id = Rising_Wedge_3p_Entry_trade_row['ID'].iloc[0]
                               # Replace this line with your API call to send a buy order
                                 
                               current_price = float(spot_prices[i])
                                                        
                               qty = int(math.floor(Total_Cash_per_position/current_price))
                                 
                               entry_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                                      
                                 
                               # ZERODHA PLACING BUY ORDER AND GET EXECUTED PRICE FROM POSITION BOOK
                               
                               ###
                               
                               kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= i, transaction_type='BUY', quantity=qty, product='CNC', order_type='MARKET')
                               
                               time.sleep(15)
                               
                               position_df = pd.DataFrame(kite.positions()["net"])
                               
                               Long_avg_price = float(position_df[position_df.tradingsymbol==i].buy_price.values[0]) 
                               
                               ###
                                                                                            
                               future_entry = float(Long_avg_price) #float(position_df[position_df.tradingsymbol==fut].buy_price.values[0])                                           
                                                      
                                  
                               # Step 4: Update Long_Trades.xlsx
                               point = Rising_Wedge_3p_Entry_trade_row['Points'].iloc[0]
                                                            
                               target_price = int(math.ceil(future_entry + (point * RISING_WEDGE_3P_LONG_targetx)))
                                                            
                               update_long_trades(trade_id, i,entry_time,Long_avg_price,qty,target_price,Rising_Wedge_3p_Entry_tradefile)
                                                             
                               # Update trading_parameters.xlsx
                               update_trade_status(Rising_Wedge_3p_Entry_tradingparameter, trade_id)
                                  
                               time.sleep(3)
                                  
                               Rising_Wedge_3p_Entry_df = pd.read_excel(Rising_Wedge_3p_Entry_tradingparameter)  
                                  
                               Rising_Wedge_3p_Entry_dft = pd.read_excel(Rising_Wedge_3p_Entry_tradefile)
                                  
                               time.sleep(3)
                                  
                               # Step 1 check Trade_Status = 'OPEN' and Action = ''     
                               Rising_Wedge_3p_Entry_valid_trades = Rising_Wedge_3p_Entry_df[(Rising_Wedge_3p_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_3p_Entry_df['Action']))]
                                                         
                               # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                               Rising_Wedge_3p_open_trades = Rising_Wedge_3p_Entry_dft[(Rising_Wedge_3p_Entry_dft['Trade Status'] == 'OPEN')]
                                                         
                               # Check Action column = 'Trade Taken'
                               Rising_Wedge_3p_taken_trades = Rising_Wedge_3p_Entry_df[Rising_Wedge_3p_Entry_df['Action'] == 'Trade Taken']
                                                            
                               tele_msg(' Rising Wedge 3 Point '+i+' '+' Long entry avg price '+str(Long_avg_price)+' Target '+str(target_price)+' And Qty '+ str(qty))
                                 

                                                      
                                           
               # Checking Target Hit
               # Check Action column = 'Trade Taken'
               Rising_Wedge_3p_taken_trades = Rising_Wedge_3p_Entry_df[Rising_Wedge_3p_Entry_df['Action'] == 'Trade Taken']

               # Loop through each taken trade
               for index, row in Rising_Wedge_3p_taken_trades.iterrows():
                   trade_id = row['ID']
                   
                   # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                   Rising_Wedge_3p_entry_trade = Rising_Wedge_3p_Entry_dft[(Rising_Wedge_3p_Entry_dft['ID'] == trade_id) & (Rising_Wedge_3p_Entry_dft['Trade Status'] == 'OPEN')]
                   
                   if not Rising_Wedge_3p_entry_trade.empty:
                       sym = Rising_Wedge_3p_entry_trade['Symbol'].iloc[0]
                       if sym == i:  # Check if the ticker matches the symbol from the Excel sheet
                           target_price = Rising_Wedge_3p_entry_trade['Target Price'].iloc[0]
                
                           # Compare spot price with stop loss price
                           if spot_prices[i] > target_price:
                               
                               qty = int(Rising_Wedge_3p_entry_trade['Qty'].iloc[0])
                               
                               fut = Rising_Wedge_3p_entry_trade['Symbol'].iloc[0]
                               
                               
                               # Send sell order to API
                               # Replace this line with your API call to send a sell order
                               
                               # Send sell order to API
                               # Replace this line with your API call to send a sell order                              

                               exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                               
                                                          
                               # ZERODHA LONG TRADE EXIT AND BOOKING PROFITS
                               
                               ###
                               
                               kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= sym, transaction_type='SELL', quantity=qty, product='CNC', order_type="LIMIT",price = target_price)
                                                    
                               time.sleep(15)
                                                         
                                   
                               position_df = pd.DataFrame(kite.positions()["net"])
                               
                               Long_exit_avg_price = float(position_df[position_df.tradingsymbol==i].sell_price.values[0])  
                               
                               ###
                               
                                                          
                               # Update Long_Trades.xlsx
                               buy_exit = Long_exit_avg_price # float(position_df[position_df.tradingsymbol==fut].sell_price.values[0])                                           
                               
                               buy_entry = float(Rising_Wedge_3p_entry_trade['Buy Price'].iloc[0])
                                                          
                               points = buy_exit - buy_entry
                               
                               broker = (((qty*buy_entry)+(qty*buy_exit)) * 0.0011)
                               
                               profit_loss = ((points * qty) - broker)
                
                               update_long_trades_target_exit(trade_id,exit_time, buy_exit, points,broker,profit_loss,Rising_Wedge_3p_Entry_tradefile)                                
                              
                
                               update_trade_status_with_trade_closed(Rising_Wedge_3p_Entry_tradingparameter, trade_id,15)
                               
                               time.sleep(3)
                               
                               Rising_Wedge_3p_Entry_df = pd.read_excel(Rising_Wedge_3p_Entry_tradingparameter)
                               
                               Rising_Wedge_3p_Entry_dft = pd.read_excel(Rising_Wedge_3p_Entry_tradefile)
                               
                               time.sleep(3)
                               
                               # Step 1 check Trade_Status = 'OPEN' and Action = ''     
                               Rising_Wedge_3p_Entry_valid_trades = Rising_Wedge_3p_Entry_df[(Rising_Wedge_3p_Entry_df['Trade_Status'] == 'OPEN') & (pd.isna(Rising_Wedge_3p_Entry_df['Action']))]                               
                               
                               # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                               Rising_Wedge_3p_open_trades = Rising_Wedge_3p_Entry_dft[(Rising_Wedge_3p_Entry_dft['Trade Status'] == 'OPEN')]                              
                               
                               # Check Action column = 'Trade Taken'
                               Rising_Wedge_3p_taken_trades = Rising_Wedge_3p_Entry_df[Rising_Wedge_3p_Entry_df['Action'] == 'Trade Taken']
                            
                            
                               tele_msg(' Rising Wedge 3 Point '+i+' Tp Hit '+ ' Sold at Price '+str(Long_exit_avg_price)+' And Profit Rupees '+str(profit_loss))

                                                              
                 
        except Exception as e:
            print("Error:",e)
            print("Oops!", e.__class__, "occurred.")
            
            ct=dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
            error_message = f"{ct} - An error occurred: {e}"
            tele_msg(error_message)
            with open("error_log.txt", "a") as error_log_file:
                error_log_file.write(error_message + "\n")
            raise ValueError("I have raised an Exception in main") 


# kite.place_order(variety ='regular', exchange='NSE', tradingsymbol= "TCS", transaction_type='BUY', quantity=1, product='CNC', order_type='MARKET')



# a = pd.DataFrame(kite.trades())

# kite.positions()

# # kite symbol only for future
# def Kite_Symbol(sym):
    
#     code = instrument_lookup(symbol=sym)
    
#     kite_sym = nse_fut[nse_fut.exchange_token == str(code)].tradingsymbol.values[0]
    
#     return kite_sym


# Kite_Symbol('BANKNIFTY 28 OCT 2025')


# type(nse_fut['exchange_token'].iloc[-1])


# kite.place_order(variety ='regular', exchange='NFO', tradingsymbol= 'BANKNIFTY25OCTFUT', transaction_type='BUY', quantity=35, product='NRML', order_type='MARKET')
# kite.place_order(variety ='regular', exchange='NFO', tradingsymbol=fut, transaction_type='BUY', quantity=qty, product='NRML', order_type='MARKET')

