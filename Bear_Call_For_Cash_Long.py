import os
working_dir = os.chdir(r"D:\ALGO TRADING\LIVE\ZERODHA CASH")
import warnings
# Suppress future warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import pytz
UTC = pytz.timezone('Asia/Kolkata')
import time


import pandas as pd
from openpyxl import load_workbook
import datetime as dt

##### Importing From User Input Library ######
from User_Input import EXIT_TIME,tele_msg,qty_dict,Hedge_Treshold,max_hedge,Tickers,lot
##############################################

##### Importing Streaming LTP DATA #############
from Streaming_Ltp import spot_prices
##############################################


##### Import From Common Functions ######
from Common_Functions import spot,first_ask,first_bid,stock_ce_option,opt_exp,data_list,is_required_time
#########################################






long_pattern_Hedge_Option_Sell_tradefile = 'Cash_long_pattern_Hedge_Option_Sell.xlsx'

long_pattern_Spread_Option_Buy_tradefile = 'Cash_long_pattern_Spread_Option_Sell.xlsx'



# Create Excel Sheet only of it Needed or first time
def create_excel_sheet(filename):
    if not os.path.exists(filename):
        columns = ['ID', 'Symbol', 'Future', 'Expiry','Entry Time','Spot Entry', 'Position Count', 'Stop Loss Price', 'Fut Buy Price', 'Qty','Exit Time', 'Fut Sell Price','Points','Brokerage','Profit/Loss', 'Trade Status','Duration(Days)']
        df = pd.DataFrame(columns=columns)
        df.to_excel(filename, index=False)
        print(f"{filename} created successfully!")
    else:
        print(f"{filename} already exists.") 


create_excel_sheet(long_pattern_Hedge_Option_Sell_tradefile) # 1

create_excel_sheet(long_pattern_Spread_Option_Buy_tradefile) # 2

# Update Long trade entry
def update_Long_trades(trade_id, ticker, fut, expiry,entry_time,spotentry, Position_Counts, sl, qty_dict,futbuy,filename):
    workbook = load_workbook(filename)
    worksheet = workbook.active

    new_row = [trade_id, ticker, fut, expiry,entry_time,spotentry, Position_Counts, sl,futbuy, qty_dict[ticker]['qty'],'','','','','', 'OPEN','']
    worksheet.append(new_row)

    workbook.save(filename)


# Update Short trade entry
def update_Short_trades(trade_id, ticker, fut, expiry,entry_time,spotentry, Position_Counts, sl, qty_dict,futsell,filename):
    workbook = load_workbook(filename)
    worksheet = workbook.active

    new_row = [trade_id, ticker, fut, expiry,entry_time,spotentry, Position_Counts, sl,'', qty_dict[ticker]['qty'],'',futsell,'','','', 'OPEN','']
    worksheet.append(new_row)

    workbook.save(filename)



# 🔹 Stop Loss Exit Update
def update_Short_trades_stop_loss_exit(trade_id, PositionCount, exit_time, buy_entry, points, brokerage, profit_loss, filename):
    wb = load_workbook(filename)
    ws = wb.active

    trade_row = None
    entry_time = None

    # Find matching row where Trade_Status is still OPEN
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        trade_id_cell = row[0].value         # Column 1: ID
        position_count_cell = row[6].value   # Column 7: Position Count
        trade_status_cell = row[15].value    # Column 16: Trade Status

        if trade_id_cell == trade_id and position_count_cell == PositionCount and trade_status_cell == "OPEN":
            trade_row = row[0].row
            entry_time = ws.cell(row=trade_row, column=5).value  # Column 5: Entry Time
            break

    if trade_row is not None and entry_time is not None:
        ws.cell(row=trade_row, column=9).value = buy_entry
        ws.cell(row=trade_row, column=11).value = exit_time
        ws.cell(row=trade_row, column=13).value = points
        ws.cell(row=trade_row, column=14).value = brokerage
        ws.cell(row=trade_row, column=15).value = profit_loss
        ws.cell(row=trade_row, column=16).value = 'Stop Loss Hit'

        try:
            entry_time_dt = dt.datetime.strptime(entry_time, "%d-%b-%Y %I:%M%p")
            exit_time_dt = dt.datetime.strptime(exit_time, "%d-%b-%Y %I:%M%p")
            ws.cell(row=trade_row, column=17).value = (exit_time_dt - entry_time_dt).days
        except Exception as e:
            print("Error parsing dates:", e)

        wb.save(filename)
    else:
        print("No OPEN trade found for Trade ID:", trade_id, "Position:", PositionCount)


# 🔹 Stop Loss Exit Update
def update_Long_trades_stop_loss_exit(trade_id, PositionCount, exit_time, sell_entry, points, brokerage, profit_loss, filename):
    wb = load_workbook(filename)
    ws = wb.active

    trade_row = None
    entry_time = None

    # Find matching row where Trade_Status is still OPEN
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        trade_id_cell = row[0].value         # Column 1: ID
        position_count_cell = row[6].value   # Column 7: Position Count
        trade_status_cell = row[15].value    # Column 16: Trade Status

        if trade_id_cell == trade_id and position_count_cell == PositionCount and trade_status_cell == "OPEN":
            trade_row = row[0].row
            entry_time = ws.cell(row=trade_row, column=5).value  # Column 5: Entry Time
            break

    if trade_row is not None and entry_time is not None:
        ws.cell(row=trade_row, column=12).value = sell_entry
        ws.cell(row=trade_row, column=11).value = exit_time
        ws.cell(row=trade_row, column=13).value = points
        ws.cell(row=trade_row, column=14).value = brokerage
        ws.cell(row=trade_row, column=15).value = profit_loss
        ws.cell(row=trade_row, column=16).value = 'Stop Loss Hit'

        try:
            entry_time_dt = dt.datetime.strptime(entry_time, "%d-%b-%Y %I:%M%p")
            exit_time_dt = dt.datetime.strptime(exit_time, "%d-%b-%Y %I:%M%p")
            ws.cell(row=trade_row, column=17).value = (exit_time_dt - entry_time_dt).days
        except Exception as e:
            print("Error parsing dates:", e)

        wb.save(filename)
    else:
        print("No OPEN trade found for Trade ID:", trade_id, "Position:", PositionCount)




# 🔹 Expiry Close Update
def update_Short_trades_expiry_close(trade_id, PositionCount, exit_time, buy_entry, points, brokerage, profit_loss, filename):
    wb = load_workbook(filename)
    ws = wb.active

    trade_row = None
    entry_time = None

    # Find matching row where Trade_Status is still OPEN
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        trade_id_cell = row[0].value
        position_count_cell = row[6].value
        trade_status_cell = row[15].value  # Column 16

        if trade_id_cell == trade_id and position_count_cell == PositionCount and trade_status_cell == "OPEN":
            trade_row = row[0].row
            entry_time = ws.cell(row=trade_row, column=5).value
            break

    if trade_row is not None and entry_time is not None:
        ws.cell(row=trade_row, column=9).value = buy_entry
        ws.cell(row=trade_row, column=11).value = exit_time
        ws.cell(row=trade_row, column=13).value = points
        ws.cell(row=trade_row, column=14).value = brokerage
        ws.cell(row=trade_row, column=15).value = profit_loss
        ws.cell(row=trade_row, column=16).value = 'Expiry Close'

        try:
            entry_time_dt = dt.datetime.strptime(entry_time, "%d-%b-%Y %I:%M%p")
            exit_time_dt = dt.datetime.strptime(exit_time, "%d-%b-%Y %I:%M%p")
            ws.cell(row=trade_row, column=17).value = (exit_time_dt - entry_time_dt).days
        except Exception as e:
            print("Error parsing dates:", e)

        wb.save(filename)
    else:
        print("No OPEN trade found for Trade ID:", trade_id, "Position:", PositionCount)


# 🔹 Expiry Close Update
def update_Long_trades_expiry_close(trade_id, PositionCount, exit_time, sell_entry, points, brokerage, profit_loss, filename):
    wb = load_workbook(filename)
    ws = wb.active

    trade_row = None
    entry_time = None

    # Find matching row where Trade_Status is still OPEN
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        trade_id_cell = row[0].value
        position_count_cell = row[6].value
        trade_status_cell = row[15].value  # Column 16

        if trade_id_cell == trade_id and position_count_cell == PositionCount and trade_status_cell == "OPEN":
            trade_row = row[0].row
            entry_time = ws.cell(row=trade_row, column=5).value
            break

    if trade_row is not None and entry_time is not None:
        ws.cell(row=trade_row, column=12).value = sell_entry
        ws.cell(row=trade_row, column=11).value = exit_time
        ws.cell(row=trade_row, column=13).value = points
        ws.cell(row=trade_row, column=14).value = brokerage
        ws.cell(row=trade_row, column=15).value = profit_loss
        ws.cell(row=trade_row, column=16).value = 'Expiry Close'

        try:
            entry_time_dt = dt.datetime.strptime(entry_time, "%d-%b-%Y %I:%M%p")
            exit_time_dt = dt.datetime.strptime(exit_time, "%d-%b-%Y %I:%M%p")
            ws.cell(row=trade_row, column=17).value = (exit_time_dt - entry_time_dt).days
        except Exception as e:
            print("Error parsing dates:", e)

        wb.save(filename)
    else:
        print("No OPEN trade found for Trade ID:", trade_id, "Position:", PositionCount)



# Merge all long Pattern Option Selling
def all_long_open_trades_file():
    try:
        # Define the specific filenames you want to merge
        file_paths = [
            'Rising_Wedge_Continuation_A_Entry_cash_Long_Trades.xlsx',
            'Rising_Wedge_Continuation_C_Entry_cash_Long_Trades.xlsx',
            'Rising_Wedge_3p_cash_Long_Trades.xlsx'
        ]

        # Load all the Excel files into a list of dataframes
        dfs = [pd.read_excel(file_path) for file_path in file_paths]

        # Merge all dataframes into one
        merged_df = pd.concat(dfs, ignore_index=True)

        # Ensure 'Entry Time' column is in datetime format
        merged_df['Entry Time'] = pd.to_datetime(merged_df['Entry Time'], errors='coerce')

        # Sort the merged dataframe based on Entry Time
        merged_df = merged_df.sort_values(by='Entry Time')
        
        merged_df = merged_df[(merged_df['Trade Status'] == 'OPEN')]

        # Save the merged and sorted dataframe to a new Excel file
        output_file_path = 'All_Long_Open_Trades.xlsx'
        #merged_df.to_excel(output_file_path, index=False)

        print(f"Merged and sorted trades saved to {output_file_path}")
    except Exception as e:
        print(f"Error in All_Long_Open_Trades: {e}")
        
    return merged_df


# Long Hedge Condition
def is_3_long_candle_reversal(candle_df):
    if len(candle_df) < 3:
        return False

    last_3 = candle_df.iloc[-3:]
    c1, c2, c3 = last_3.iloc[0], last_3.iloc[1], last_3.iloc[2]

    cond1 = (c1['Close'] > c1['Open']) and (c1['Close'] > c2['Close']) and (c1['Close'] > c3['Close'])
    cond2 = (c2['Close'] < c2['Open']) and (c2['Close'] < c1['Close']) and (c2['Close'] > c3['Close'])
    cond3 = (c3['Close'] < c3['Open']) and (c3['Close'] < c1['Close']) and (c3['Close'] < c2['Close'])

    return cond1 and cond2 and cond3

        
# Checking Long Patterns for Hedge
def long_hedge():
    
    Long_Open_Trades = all_long_open_trades_file()
    
    hedge_trade = pd.read_excel(long_pattern_Hedge_Option_Sell_tradefile)
    
    
    for long_open in Tickers:
                
        # 1) For Long 
        Long_Open_Trades_trade_row = Long_Open_Trades[Long_Open_Trades['Symbol'] == long_open]
        Long_Open_Trades_Entry_trade_row = Long_Open_Trades_trade_row.sort_values(by="Buy Price", ascending=True)
        
        
        # Checking Entry For Long Condition for placing Hedge
        if not Long_Open_Trades_Entry_trade_row.empty:
            
            # Loop through each open trade row (multiple trades per symbol)
            for index, row in Long_Open_Trades_Entry_trade_row.iterrows():
                trade_id = row['ID']
                entry_point = row['Buy Price']
                
                #print(entry_point)
             
                
            
                # Step 1: Compare spot price with entry point
                if spot_prices[long_open] < entry_point - (entry_point * Hedge_Treshold):
                    
                    ohlc_data = data_list.get(long_open)
                    
                    if is_3_long_candle_reversal(ohlc_data):
                        
                        print(f"✅ Bearish 3-candle reversal detected for {long_open}. Consider hedging.")
                                                           
                        
                        taken_trades = hedge_trade[(hedge_trade['ID'] == trade_id) & (hedge_trade['Trade Status'] == 'OPEN')]

                        if len(taken_trades) < max_hedge:
                            
                            exp = opt_exp(long_open)
                            
                            Hedge_Option,Hedge_Option_SL,Spread_Option = stock_ce_option(long_open)                        
                            
                            qty = qty_dict[long_open]['qty']
                            
                            position_count = len(taken_trades)+1
                            
                            ###############################
                            ### Spread Option Buy Entry ###
                            ###############################
                            
                            entry_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                            
                            # Sending Orders to the API
                            # Long Entry Adding Spread
                            
                            long_entry = first_ask(Spread_Option)
                            
                            
                            # Update Option Spread trade book
                            
                            update_Long_trades(trade_id, long_open, Spread_Option, exp,entry_time,spot_prices[long_open], position_count, Hedge_Option_SL, qty_dict,long_entry,long_pattern_Spread_Option_Buy_tradefile)
                            
                            tele_msg(f'🛡️ Spread Trade Added {Spread_Option} | Spot: {spot_prices[long_open]} | SL: {Hedge_Option_SL} | Bought at: {long_entry}')
                            
                            
                            
                            
                            ###############################
                            ### Hedge Option Sell Entry ###
                            ###############################
                            
                            entry_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                            
                            
                            # Sending Orders to the API
                            short_entry = first_bid(Hedge_Option)
                            
                                                        
                            # Update Option Hedges trade book
                            update_Short_trades(trade_id, long_open, Hedge_Option, exp,entry_time,spot_prices[long_open], position_count, Hedge_Option_SL, qty_dict,short_entry,long_pattern_Hedge_Option_Sell_tradefile)
                            
                            #tele_msg(' Hedge Trade Added '+call_option+' '+' Short entry and Spot Entry '+str(spot_prices[long_open])+' Stop Loss '+str(call_stop_loss)+' And Sold at '+str(short_entry))
                            tele_msg(f'🛡️ Hedge Trade Added {Hedge_Option} | Spot: {spot_prices[long_open]} | SL: {Hedge_Option_SL} | Sold at: {short_entry}')
                            
                        else:
                            
                            print("Maximum Hedge Count Reached ")


def Cash_Bear_Call_Spread():
    
    Long_Open_Trades = all_long_open_trades_file()
    
    hedge_trade = pd.read_excel(long_pattern_Hedge_Option_Sell_tradefile)
                           
    # Check Action column = 'Trade Taken'
    long_hedge_trades_taken_trades = hedge_trade[hedge_trade['Trade Status'] == 'OPEN']
   
    
    
    endTime = dt.datetime.now(pytz.timezone('Asia/Kolkata')).replace(hour=EXIT_TIME[0], minute=EXIT_TIME[1],second=EXIT_TIME[2])
    while dt.datetime.now(pytz.timezone('Asia/Kolkata')) <  endTime:

        try:

            for i in Tickers:
                
                # spot_prices[i] = spot(i)


                # print("#################################################")
                # print("### THIS IS HEDGE TRADE FOR CASH STOCK TRADES ###")
                # print("#################################################")

                # print('Spot prices of',i,' ',spot_prices[i])
                
                
                # Updating Candles only on a specific time
                if is_required_time():
                    
                    time.sleep(120)
                    
                    long_hedge()

                    # Checking Long Hedge Trades
                    hedge_trade = pd.read_excel(long_pattern_Hedge_Option_Sell_tradefile)
                                       
                    # Check Action column = 'Trade Taken'
                    long_hedge_trades_taken_trades = hedge_trade[hedge_trade['Trade Status'] == 'OPEN']
                    
                
            
                #######################
                #######################
                ### For Long Hedge ####
                #######################
                #######################
                
                # Checking Stop Loss for Long option selling
                
                # Loop through each taken trade
                for index, row in long_hedge_trades_taken_trades.iterrows():
                    
                    sym = row['Symbol']
                    
                    if sym == i:
                        
                        PositionCount = row['Position Count']
                        
                        trade_id = row['ID'] 
                        
                        long_hedge_trades_taken = long_hedge_trades_taken_trades[(long_hedge_trades_taken_trades['Symbol'] == sym) & (long_hedge_trades_taken_trades['Trade Status'] == 'OPEN')]
                        
                        long_hedge_trades_taken = long_hedge_trades_taken_trades[(long_hedge_trades_taken_trades['Position Count'] == PositionCount) & (long_hedge_trades_taken_trades['ID'] == trade_id)]
                        
                                                                                         
                        # Check current position
                        if not long_hedge_trades_taken.empty:
                             
                            print("cheking for long hedge sl")
                            print(sym)
                      
                            stop_loss_price = row['Stop Loss Price']
                  
                            print(stop_loss_price)
                             
                            # Check spot price is above stop loss
                            if spot_prices[i] > stop_loss_price:
                                 
                                qty = row['Qty']
                                 
                                Hedge_Option = row['Future']
                                
                                
                                ###############################
                                ### Hedge Option Sell Exit  ###
                                ###############################
                                
                                 
                                                                
                                # Send sell order to API
                                # Replace this line with your API call to send a sell order                              
                                 
                                exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                 
                                Hedge_Exit = first_ask(Hedge_Option)                
                                 
                                # Update Long_Trades.xlsx
                                  
                                Hedge_Entry = float(row['Fut Sell Price'])
                                 
                                points =  Hedge_Entry - Hedge_Exit
                                 
                                broker = lot * (((qty*Hedge_Entry)+(qty*Hedge_Exit)) * 0.0045)
                                 
                                profit_loss = lot * ((points * qty) - broker)
                                 
                                update_Short_trades_stop_loss_exit(trade_id,PositionCount,exit_time,Hedge_Exit, points,broker, profit_loss,long_pattern_Hedge_Option_Sell_tradefile)
                                 
                                # Checking Long Hedge Trades
                                long_hedge_trades = pd.read_excel(long_pattern_Hedge_Option_Sell_tradefile)
                                # Check Action column = 'Trade Taken'
                                long_hedge_trades_taken_trades = long_hedge_trades[long_hedge_trades['Trade Status'] == 'OPEN']
          
                                tele_msg(f'🛡️ Stop loss Hit  {Hedge_Option} | Spot: {spot_prices[i]} | Bought at: {Hedge_Exit} | Hedge profit/loss: {profit_loss}')
                                
                                
                                
                                ###############################
                                ### Spread Option Buy Exit  ###
                                ###############################
                                

                                long_spread_trades = pd.read_excel(long_pattern_Spread_Option_Buy_tradefile)

                                # Find the option trade in Option Trades.xlsx with Trade Status = 'OPEN'
                                spread_option_trade = long_spread_trades[(long_spread_trades['Symbol'] == sym) & (long_spread_trades['Trade Status'] == 'OPEN')]
                                
                                spread_option_trade = long_spread_trades[(long_spread_trades['Position Count'] == PositionCount) & (long_spread_trades['ID'] == trade_id)]
                                
                                
                                if not spread_option_trade.empty:

                                    Spread_Option = spread_option_trade['Future'].iloc[0]
                                    
                                    
                                    exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                    
                                    Spread_Exit = first_ask(Spread_Option)
                                    
                                    # Update Long_Trades.xlsx
                                   
                                    Spread_Entry = float(spread_option_trade['Fut Buy Price'].iloc[0])

                                    points = Spread_Exit - Spread_Entry

                                    broker = lot * (((qty*Spread_Entry)+(qty*Spread_Exit)) * 0.0090)

                                    profit_loss = lot * ((points * qty) - broker)
                                    
                                    update_Long_trades_stop_loss_exit(trade_id, PositionCount, exit_time, Spread_Exit, points, broker, profit_loss, long_pattern_Spread_Option_Buy_tradefile)

                                    tele_msg(f'🛡️ Stop loss Hit  {Spread_Option} | Spot: {spot_prices[i]} | Bought at: {Spread_Exit} | Spread profit/loss: {profit_loss}')
                                    
                                
                # Closing Current open position at the expiry       
                       
                
                # Iterate through each row of the DataFrame
                for index, row in long_hedge_trades_taken_trades.iterrows():
                    
                    # Checking the symbol name and ticker name are matching
                    sym = row['Symbol']
                    
                    #print(symbol)
                    
                    if sym == i:
                        
                        trade_id = row['ID']
                        
                        PositionCount = row['Position Count']
                        
                        # Find the trade in Long_Trades.xlsx with Trade Status = 'OPEN'
                        
                        long_hedge_trades_taken = long_hedge_trades_taken_trades[(long_hedge_trades_taken_trades['Symbol'] == sym) & (long_hedge_trades_taken_trades['Trade Status'] == 'OPEN')]
                        
                        long_hedge_trades_taken = long_hedge_trades_taken[(long_hedge_trades_taken['Position Count'] == PositionCount) & (long_hedge_trades_taken['ID'] == trade_id)]
                         
                        
                        print("CHECKING FOR EXPIRY : ",sym)
                      
                     
                        # Check if the expiry date is within 5 days from today
                        if row['Expiry'] and (pd.to_datetime(row['Expiry']) - dt.timedelta(days=0)).date() <= dt.datetime.now().date() and dt.datetime.now().time() >= dt.time(15, 20):
                            # If the position is "OPEN", execute the rollover logic
                            if row['Trade Status'] == 'OPEN':
                                 
                                 
                                # Store the values in separate variables
                                                          
                                sl = row['Stop Loss Price']
                                qty = row['Qty']
                                spot_entry = row['Spot Entry']                               
                                Hedge_Option = row['Future']                         
                                                                                             
                                 
                                # Send sell order to API
                                # Replace this line with your API call to send a sell order 
                                
                                ###############################
                                ### Hedge Option Sell Exit  ###
                                ###############################
                                
      
                                exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                
                                Hedge_Exit = first_ask(Hedge_Option)
                                 
                                # Update Long_Trades.xlsx
                                   
                                Hedge_Entry = float(row['Fut Sell Price'])
                                 
                                points =  Hedge_Entry - Hedge_Exit 
                                 
                                broker = lot * (((qty*Hedge_Entry)+(qty*Hedge_Exit)) * 0.0045)
                                 
                                profit_loss = lot * ((points * qty) - broker)
                                 
                                time.sleep(3)   
                                                           
                                update_Short_trades_expiry_close(trade_id, PositionCount, exit_time, Hedge_Exit, points, broker, profit_loss, long_pattern_Hedge_Option_Sell_tradefile)
                                  
                                # Checking Long Hedge Trades
                                long_hedge_trades = pd.read_excel(long_pattern_Hedge_Option_Sell_tradefile)
                                # Check Action column = 'Trade Taken'
                                long_hedge_trades_taken_trades = long_hedge_trades[long_hedge_trades['Trade Status'] == 'OPEN']
      
                                tele_msg(f'🛡️ Expiry Close  {Hedge_Option} | Spot: {spot_prices[i]} | Bought at: {Hedge_Exit} | Hedge profit/loss: {profit_loss}')
                                
                                
                                ###############################
                                ### Spread Option Buy Exit  ###
                                ###############################
                                

                                long_spread_trades = pd.read_excel(long_pattern_Spread_Option_Buy_tradefile)

                                # Find the option trade in Option Trades.xlsx with Trade Status = 'OPEN'
                                spread_option_trade = long_spread_trades[(long_spread_trades['Symbol'] == sym) & (long_spread_trades['Trade Status'] == 'OPEN')]
                                
                                spread_option_trade = long_spread_trades[(long_spread_trades['Position Count'] == PositionCount) & (long_spread_trades['ID'] == trade_id)]
                                
                                
                                if not spread_option_trade.empty:

                                    Spread_Option = spread_option_trade['Future'].iloc[0]
                                    
                                    
                                    exit_time = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
                                    
                                    Spread_Exit = first_ask(Spread_Option)
                                    
                                    # Update Long_Trades.xlsx
                                   
                                    Spread_Entry = float(spread_option_trade['Fut Buy Price'].iloc[0])

                                    points = Spread_Exit - Spread_Entry

                                    broker = lot * (((qty*Spread_Entry)+(qty*Spread_Exit)) * 0.0090)

                                    profit_loss = lot * ((points * qty) - broker)
                                    
                                    update_Long_trades_stop_loss_exit(trade_id, PositionCount, exit_time, Spread_Exit, points, broker, profit_loss, long_pattern_Spread_Option_Buy_tradefile)

                                    tele_msg(f'🛡️ Stop loss Hit  {Spread_Option} | Spot: {spot_prices[i]} | Bought at: {Spread_Exit} | Spread profit/loss: {profit_loss}')
                                    
                                
                                
                                
                                
      
                    
                    
               
        except Exception as e:
            print("Error:",e)
            print("Oops!", e.__class__, "occurred.")
            
            ct=dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
            error_message = f"{ct} - An error occurred: {e}"
            tele_msg(error_message)
            with open("error_log.txt", "a") as error_log_file:
                error_log_file.write(error_message + "\n")
            raise ValueError("I have raised an Exception in main") 


    algostop = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")

    print("Algo stopped Time : ",algostop)            
























    
    

